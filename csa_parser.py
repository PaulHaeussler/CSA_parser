import json
import os
import warnings
import sys
import pywintypes
import win32com.client
import openpyxl
from colorama import Fore, Back, Style, init
from openpyxl.reader.excel import load_workbook

DEBUG = False
DEBUG_PATH = "C:\\Users\\U1282433\\Downloads\\digital rewe alt.xlsx"
COMMENT = ""

init(convert=True)

def print_title():
    print(Fore.LIGHTCYAN_EX)
    print("   _____  _____                                        ")
    print("  / ____|/ ____|  /\                                   ")
    print(" | |    | (___   /  \   _ __   __ _ _ __ ___  ___ _ __ ")
    print(" | |     \___ \ / /\ \ | '_ \ / _` | '__/ __|/ _ \ '__|")
    print(" | |____ ____) / ____ \| |_) | (_| | |  \__ \  __/ |   ")
    print("  \_____|_____/_/    \_\ .__/ \__,_|_|  |___/\___|_|   ")
    print("                       | |                             ")
    print("                       |_|                             ")
    print(Style.RESET_ALL)


def isN(val):
    if val is None:
        return True
    elif val == "":
        return True
    else:
        return False


class Question:
    pass


class Subsection:
    pass


class Section:

    def __init__(self, ws, csa):
        self.ws = ws
        print(f"{Fore.LIGHTCYAN_EX} Parsing {self.ws.title}{Style.RESET_ALL}")

        self.sheet_name = self.ws.title               # worksheet name '1 Demographics'
        self.index = self.ws.title.partition(" ")[0]  # index 1
        self.split_sheet_name = self.ws.title.partition(" ")[2]   # only the name 'Demographics'
        self.name = self.ws.cell(1, 1).value
        self.subsections = []
        self.parse_section(csa)
        print()

    def testMerge(self, row, column):
        cell = self.ws.cell(row, column)
        for mergedCell in self.ws.merged_cells.ranges:
            if cell.coordinate in mergedCell:
                return True
        return False


    def readValidations(self):
        result = {}
        for dv in self.ws.data_validations.dataValidation:
            frm = dv.formula1
            if frm in result.keys():
                continue
            vals = []
            for row in self.ws[frm]:
                vals.append(row[0].value)
            result[frm] = vals
        return result

    def isDropdown(self, cell):
        for v in self.ws.data_validations.dataValidation:
            tmp = list(v.ranges.ranges)[0].top[0]              # bisschen umständlich, vergleiche die koordinaten von den cells mit validations (dropdowns) gegen aktuelle zelle. Annahme: alle validations beziehen sich auf eine range bestehend aus einer einzelnen zelle (deshalb ist es egal ob wir top bot right or left nehmen)
            if cell.row == tmp[0] and cell.column == tmp[1]:
                return self.validations[v.formula1]
        return None

    def parse_section(self, csa):
        row = 4
        current_ss = None
        q = None
        self.validations = self.readValidations()

        while True:
            a = self.ws.cell(row=row, column=1).value
            b = self.ws.cell(row=row, column=2).value
            c = self.ws.cell(row=row, column=3).value
            d = self.ws.cell(row=row, column=4).value
            a_plus_1 = self.ws.cell(row=row + 1, column=1).value
            b_plus_1 = self.ws.cell(row=row + 1, column=2).value
            c_plus_1 = self.ws.cell(row=row + 1, column=3).value
            d_plus_1 = self.ws.cell(row=row + 1, column=4).value
            if isN(a) and isN(b) and isN(c):  # no values in columns A, B and C means end of sheet
                break

            # parse subsection header
            if self.testMerge(row, 1):
                ss = Subsection()
                ss.name = a
                ss.questions = []
                self.subsections.append(ss)
                current_ss = ss
            # parse question, muss frage sein wenn inhalt in A und kein section header
            elif not isN(a):
                print(f"    {a} {current_ss.name} {b} ")
                q = Question()
                q.index = a
                q.question = b
                q.isAnswered = False
                current_ss.questions.append(q)
                if not isN(a_plus_1) or not isN(b_plus_1) or (isN(a_plus_1) and isN(b_plus_1) and isN(c_plus_1)):   # da keine weiteren Zeilen (nächste Frage/Sectionheader direkt darunter oder EOF darunter) ist keine single/multiple choice
                    vals = self.isDropdown(self.ws.cell(row=row, column=4))
                    if d is None:
                        q.answer = ""
                    else:
                        q.answer = d
                        q.isAnswered = True
                    if vals is None:  # keine optionen im dropdown / keine data validation bedeutet freitext
                        q.type = "T"   # text
                    elif vals == ['Yes', 'No']:  # yes / no binary question
                        q.type = "B"  # binary
                    else:     # es gibt dropdown, aber mit anderen werten als yes/no (z.b. länder)
                        q.type = "S"   # select
                        q.options = {}
                        for v in vals:
                            q.options[v] = v == q.answer
                elif isN(a_plus_1) and isN(b_plus_1) and not isN(c_plus_1):    # multiple choice, sichergehen
                    row_offset = 1
                    q.options = {}
                    q.type = "M"
                    while row_offset < 50:     # Annahme: keine felder mit mehr als 50 optionen (failsave um forever while loop zu vermeiden)
                        if (not isN(self.ws.cell(row=row + row_offset, column=1).value) or not isN(self.ws.cell(row=row + row_offset, column=2).value)) or isN(self.ws.cell(row+row_offset, 3).value):  # break if either col A or B becomes not null (new question or header) or if col C is empty (EOF)
                            break
                        val = False
                        if self.ws.cell(row=row + row_offset, column=4).value == "Yes":
                            val = True
                            q.isAnswered = True
                        q.options[self.ws.cell(row=row + row_offset, column=3).value] = val
                        row_offset += 1
                    else:
                        raise Exception(f"Failed parsing multichoice question {q.index} {q.question}")
                else:
                    raise Exception(f"This shouldnt happen, parsing error at {q.index} {q.question}")

            if a == "1.4.7":  # Saudumme Ausnahme nur bei 1.4.7 ist for some reason die Option in B statt in C
                q.options = {}
                q.type = "E"
                row_offset = 1
                while row_offset < 50:  # same as above
                    if (isN(self.ws.cell(row+row_offset, 1).value) and isN(self.ws.cell(row+row_offset, 2).value) and isN(self.ws.cell(row+row_offset, 3).value) and isN(self.ws.cell(row+row_offset, 4).value)) or self.ws.cell(row+row_offset, 2).value == COMMENT:
                        break
                    v = self.ws.cell(row+row_offset, 4).value
                    if v is None:
                        v = ""
                    q.options[self.ws.cell(row+row_offset, 2).value] = v
                    row_offset += 1

            if b == COMMENT and d is not None:
                q.comment = d
                q.isAnswered = True

            if not isN(d):
                if "??" in d or "tbv" in d:
                    csa.tbvs.append(q)



            row += 1
class CSA:

    def __init__(self):
        try:
            self.x = win32com.client.GetActiveObject("Excel.Application")
        except pywintypes.com_error:
            raise Exception("Die Input-Exceldatei bitte zuvor öffnen!")
        print_title()
        self.origin_file = self.select_wb()
        with warnings.catch_warnings(record=True):  # Suppressing irrelevant warning here
            warnings.simplefilter("always")
            self.wb = load_workbook(self.origin_file)
        self.language = self.check_lang()
        self.sections = []
        self.tbvs = []
        del self.x
        for ws in self.wb.worksheets:
            if not ws.title in ['Details', 'Einzelheiten']:
                self.sections.append(Section(ws, self))
        self.check_unanswered()
        # printing tbvs
        self.print_tbvs()
        print("Sucessfully parsed CSA!")


    def check_unanswered(self):
        unanswered = []
        answered = []
        for s in self.sections:
            for ss in s.subsections:
                for q in ss.questions:
                    if q.isAnswered:
                        answered.append(q)
                    else:
                        unanswered.append(q)
        total_length = len(unanswered) + len(answered)
        self.question_count = total_length
        self.section_count = len(self.sections)
        self.questions_answered = len(answered)
        self.questions_unanswered = len(unanswered)
        self.ss_count = 0
        for s in self.sections:
            for ss in s.subsections:
                self.ss_count += 1
        print(f"{Fore.LIGHTGREEN_EX}{len(answered)} out of {len(answered) + len(unanswered)} were answered{Style.RESET_ALL}")
        if 0 < len(unanswered) < 10:  # print questions in question if theres only a handful of em
            print(f"{Fore.LIGHTRED_EX}Found unanswered questions! See below:{Style.RESET_ALL}")
            for q in unanswered:
                print(f"  {Fore.LIGHTYELLOW_EX}{q.index} {q.question}{Style.RESET_ALL}")
        if len(unanswered) > 0 and not DEBUG:
            i = input(f"{Fore.LIGHTRED_EX}Found {len(unanswered)} unanswered questions! Type 'Y' to continue and any other key to print them and abort:{Style.RESET_ALL}")
            if i.upper() == "Y":
                print("Continuing...")
            else:
                for q in unanswered:
                    print(f"  {q.index} {q.question}")
                sys.exit("Exiting...")

    def print_tbvs(self):
        if len(self.tbvs) > 0:
            # a lot of code just to markup the relevant part
            for tbv in self.tbvs:
                tmp = None
                if tbv.type == "M":
                    if "??" in tbv.comment:
                        tmp = tbv.comment.partition("??")
                    else:  # tbv
                        tmp = tbv.comment.partition("tbv")
                elif "??" in tbv.answer:
                    tmp = tbv.answer.partition("??")
                elif "tbv" in tbv.answer:
                    tmp = tbv.answer.partition("tbv")
                else:
                    if hasattr(tbv, "comment"):
                        if "??" in tbv.comment:
                            tmp = tbv.comment.partition("??")
                        elif "tbv" in tbv.comment:
                            tmp = tbv.comment.partition("tbv")
                print(f"  {Fore.LIGHTCYAN_EX}{tbv.index}{Style.RESET_ALL} {tmp[0]}{Fore.RED}{tmp[1]}{Style.RESET_ALL}{tmp[2]}")
            if not DEBUG:
                i = input(
                    Fore.RED + Style.BRIGHT + Back.BLACK + "WARNING: Unresolved TBVs found. Type 'Y' to continue and any other key to abort:" + Style.RESET_ALL)
                if i.upper() == "Y":
                    print("Continuing...")
                else:
                    sys.exit("Exiting...")

    def select_wb(self):
        workbooks = []
        for wb in self.x.Workbooks:
            workbooks.append(wb.FullName)
            print(f"{len(workbooks)}. {wb.fullName}")
            del wb
        if len(self.x.Workbooks) > 0:
            if DEBUG:
                return workbooks[0]
            while True:
                inpt = input("Type the number of the opened workbook to use as input:")
                if not inpt.isnumeric():
                    print("Invalid entry, try again")
                    continue
                i = int(inpt)
                if i > 0 and i <= len(workbooks):
                    return workbooks[i-1]
                else:
                    print("Invalid workbook, try again")
        else: 
            if DEBUG:
                return DEBUG_PATH
            while True:
                i = input("Bitte Pfad zur Mappe angeben: ")
                if os.path.isfile(i):
                    return i


    def check_lang(self):
        for ws in self.wb.worksheets:
            if ws.title[:2] == "1 ":
                global COMMENT
                if ws.title[2:] == "Demographics":
                    COMMENT = "Comment"
                    return "EN"
                elif ws.title[2:] == "Demographische Daten":
                    COMMENT = "Comment"  #TODO: Anpassen sobald die keks auch mal das im Export korrigieren
                    return "DE"
                else:
                    raise Exception(f"Demographics sheet in unknown language! {ws.title}")
            del ws
        raise Exception("Couldn't find any sheet starting with '1 ', is this really a CSA Export?")


if __name__ == "__main__":
    csa = CSA()
